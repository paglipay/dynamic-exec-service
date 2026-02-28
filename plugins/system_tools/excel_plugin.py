"""Excel plugin for exporting sheet data to JSON."""

from __future__ import annotations

import json
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook


class ExcelPlugin:
    """Read Excel sheets and save selected data as JSON."""

    def __init__(self, base_dir: str = "generated_data", allow_outside_base_dir: bool = True) -> None:
        if not isinstance(base_dir, str) or not base_dir.strip():
            raise ValueError("base_dir must be a non-empty string")
        if not isinstance(allow_outside_base_dir, bool):
            raise ValueError("allow_outside_base_dir must be a boolean")

        self.base_dir = Path(base_dir).resolve()
        self.allow_outside_base_dir = allow_outside_base_dir
        if not self.base_dir.exists() or not self.base_dir.is_dir():
            raise ValueError("base_dir must point to an existing directory")

    def _resolve_path(self, path_value: str) -> Path:
        if not isinstance(path_value, str) or not path_value.strip():
            raise ValueError("path must be a non-empty string")
        raw = Path(path_value.strip())
        resolved = raw.resolve() if raw.is_absolute() else (self.base_dir / raw).resolve()
        if not self.allow_outside_base_dir:
            try:
                resolved.relative_to(self.base_dir)
            except ValueError as exc:
                raise ValueError("path must be inside base_dir") from exc
        return resolved

    def _apply_filters(self, frame: pd.DataFrame, filter_by: list[dict[str, Any]]) -> pd.DataFrame:
        filtered = frame
        for item in filter_by:
            if not isinstance(item, dict):
                raise ValueError("Each filter_by entry must be an object")

            column = item.get("column")
            operator = item.get("operator", "equals")
            value = item.get("value")

            if not isinstance(column, str) or not column:
                raise ValueError("filter column must be a non-empty string")
            if column not in filtered.columns:
                raise ValueError(f"filter column not found: {column}")
            if not isinstance(operator, str) or not operator:
                raise ValueError("filter operator must be a non-empty string")

            op = operator.lower()
            series = filtered[column]

            if op in {"equals", "eq", "=="}:
                filtered = filtered[series == value]
            elif op in {"not_equals", "neq", "!="}:
                filtered = filtered[series != value]
            elif op in {"contains"}:
                filtered = filtered[series.astype(str).str.contains(str(value), case=False, na=False)]
            elif op in {"in"}:
                if not isinstance(value, list):
                    raise ValueError("'in' filter value must be an array")
                filtered = filtered[series.isin(value)]
            else:
                raise ValueError(f"unsupported filter operator: {operator}")

        return filtered

    def _to_json_safe(self, value: Any) -> Any:
        if isinstance(value, dict):
            return {key: self._to_json_safe(item) for key, item in value.items()}
        if isinstance(value, (list, tuple)):
            return [self._to_json_safe(item) for item in value]
        if value is None:
            return None

        try:
            if pd.isna(value):
                return None
        except Exception:
            pass

        if hasattr(value, "isoformat") and callable(getattr(value, "isoformat")):
            try:
                return value.isoformat()
            except Exception:
                pass

        if hasattr(value, "item") and callable(getattr(value, "item")):
            try:
                return value.item()
            except Exception:
                pass

        return value

    def excel_to_json(
        self,
        file_path: str,
        sheet: str | int = 0,
        columns: list[str] | None = None,
        filter_by: list[dict[str, Any]] | None = None,
        save_as: str = "output.json",
    ) -> dict[str, Any]:
        """Read an Excel sheet, optionally select/filter rows, and save to a JSON file."""
        excel_path = self._resolve_path(file_path)
        if not excel_path.exists() or not excel_path.is_file():
            raise ValueError("file_path does not exist")
        if excel_path.suffix.lower() not in {".xlsx", ".xlsm", ".xls"}:
            raise ValueError("file_path must be an Excel file (.xlsx/.xlsm/.xls)")

        if not isinstance(sheet, (str, int)):
            raise ValueError("sheet must be a string or integer")

        if columns is not None:
            if not isinstance(columns, list) or any(not isinstance(item, str) or not item for item in columns):
                raise ValueError("columns must be an array of non-empty strings when provided")

        if filter_by is not None and not isinstance(filter_by, list):
            raise ValueError("filter_by must be an array when provided")

        output_path = self._resolve_path(save_as)
        if output_path.suffix.lower() != ".json":
            raise ValueError("save_as must be a .json file")

        try:
            frame = pd.read_excel(excel_path, sheet_name=sheet)
        except Exception as exc:
            raise ValueError(f"Failed to read Excel file: {exc}") from exc

        if columns:
            missing_columns = [column for column in columns if column not in frame.columns]
            if missing_columns:
                raise ValueError(f"columns not found in sheet: {', '.join(missing_columns)}")
            frame = frame[columns]

        if filter_by:
            frame = self._apply_filters(frame, filter_by)

        records = self._to_json_safe(frame.where(pd.notna(frame), None).to_dict(orient="records"))

        output_path.parent.mkdir(parents=True, exist_ok=True)
        try:
            output_path.write_text(json.dumps(records, ensure_ascii=False, indent=2), encoding="utf-8")
        except Exception as exc:
            raise ValueError(f"Failed to write JSON output: {exc}") from exc

        return {
            "status": "success",
            "action": "excel_to_json",
            "file_path": str(excel_path),
            "sheet": sheet,
            "column_names": [str(column) for column in frame.columns.tolist()],
            "selected_columns": columns if columns is not None else "all",
            "filters_applied": len(filter_by) if isinstance(filter_by, list) else 0,
            "row_count": len(records),
            "save_as": str(output_path),
        }

    def list_columns_in_sheet(self, file_path: str, sheet: str | int = 0) -> dict[str, Any]:
        """Return column metadata for a single sheet in a workbook."""
        excel_path = self._resolve_path(file_path)
        if not excel_path.exists() or not excel_path.is_file():
            raise ValueError("file_path does not exist")
        if excel_path.suffix.lower() not in {".xlsx", ".xlsm", ".xls"}:
            raise ValueError("file_path must be an Excel file (.xlsx/.xlsm/.xls)")
        if not isinstance(sheet, (str, int)):
            raise ValueError("sheet must be a string or integer")

        try:
            workbook = pd.ExcelFile(excel_path)
        except Exception as exc:
            raise ValueError(f"Failed to open Excel file: {exc}") from exc

        if isinstance(sheet, int):
            if sheet < 0 or sheet >= len(workbook.sheet_names):
                raise ValueError(f"sheet index out of range: {sheet}")
            sheet_index = sheet
            sheet_name = workbook.sheet_names[sheet]
        else:
            if sheet not in workbook.sheet_names:
                raise ValueError(f"sheet not found: {sheet}")
            sheet_name = sheet
            sheet_index = workbook.sheet_names.index(sheet)

        try:
            frame = pd.read_excel(excel_path, sheet_name=sheet_name)
        except Exception as exc:
            raise ValueError(f"Failed to read sheet '{sheet_name}': {exc}") from exc

        column_names = [str(column) for column in frame.columns.tolist()]
        first_data_row = self._to_json_safe(frame.iloc[0].tolist()) if not frame.empty else []

        return self._to_json_safe({
            "status": "success",
            "action": "list_columns_in_sheet",
            "file_path": str(excel_path),
            "sheet_index": sheet_index,
            "sheet_name": sheet_name,
            "row_count": int(len(frame)),
            "column_count": int(len(column_names)),
            "column_names": column_names,
            "first_row_column_names": column_names,
            "first_data_row": first_data_row,
        })

    def list_sheet_names(self, file_path: str) -> dict[str, Any]:
        """Return all sheet names for a workbook."""
        excel_path = self._resolve_path(file_path)
        if not excel_path.exists() or not excel_path.is_file():
            raise ValueError("file_path does not exist")
        if excel_path.suffix.lower() not in {".xlsx", ".xlsm", ".xls"}:
            raise ValueError("file_path must be an Excel file (.xlsx/.xlsm/.xls)")

        try:
            workbook = pd.ExcelFile(excel_path)
        except Exception as exc:
            raise ValueError(f"Failed to open Excel file: {exc}") from exc

        sheet_names = [str(name) for name in workbook.sheet_names]
        return {
            "status": "success",
            "action": "list_sheet_names",
            "file_path": str(excel_path),
            "sheet_count": int(len(sheet_names)),
            "sheet_names": sheet_names,
        }

    def append_mapped_output_change(
        self,
        output: Any,
        output_col: str,
        row: int,
        target_sheet: str,
        sheet_column_idx: dict[str, dict[str, int]],
        changes_array: list[dict[str, Any]] | None = None,
    ) -> dict[str, Any]:
        """Store output using mapped output column metadata and return updated changes array."""
        if not isinstance(output_col, str) or not output_col.strip():
            raise ValueError("output_col must be a non-empty string")
        if not isinstance(row, int) or row < 0:
            raise ValueError("row must be a non-negative integer")
        if not isinstance(target_sheet, str) or not target_sheet.strip():
            raise ValueError("target_sheet must be a non-empty string")
        if not isinstance(sheet_column_idx, dict):
            raise ValueError("sheet_column_idx must be an object")
        if changes_array is None:
            changes_array = []
        if not isinstance(changes_array, list):
            raise ValueError("changes_array must be an array when provided")

        sheet_name = target_sheet.strip()
        col_name = output_col.strip()

        if sheet_name not in sheet_column_idx:
            raise ValueError(f"target_sheet not found in sheet_column_idx: {sheet_name}")
        sheet_mapping = sheet_column_idx[sheet_name]
        if not isinstance(sheet_mapping, dict):
            raise ValueError(f"sheet_column_idx entry for '{sheet_name}' must be an object")
        if col_name not in sheet_mapping:
            raise ValueError(f"output_col not mapped for sheet '{sheet_name}': {col_name}")

        mapped_col = sheet_mapping[col_name]
        if not isinstance(mapped_col, int):
            raise ValueError("mapped column index must be an integer")

        change_item = {
            "output": self._to_json_safe(output),
            "col": mapped_col,
            "row": row,
            "sheet_name": sheet_name,
        }
        changes_array.append(change_item)

        return {
            "status": "success",
            "action": "append_mapped_output_change",
            "change": change_item,
            "changes_array": self._to_json_safe(changes_array),
            "change_count": len(changes_array),
        }

    def update_sheet_row_values(
        self,
        file_path: str | dict[str, Any],
        sheet: str | int | None = None,
        row: int | None = None,
        columns: list[str] | None = None,
        values: Any = None,
        header_row: int = 1,
        save_as: str | None = None,
    ) -> dict[str, Any]:
        """Update one worksheet row using header column names and value(s)."""
        if isinstance(file_path, dict):
            payload = file_path
            resolved_file_path = payload.get("file_path")
            resolved_sheet = payload.get("sheet", sheet)
            resolved_row = payload.get("row", row)
            resolved_columns = payload.get("columns", columns)
            resolved_values = payload.get("values", values)
            resolved_updates = payload.get("updates")
            resolved_header_row = payload.get("header_row", header_row)
            resolved_save_as = payload.get("save_as", save_as)
        else:
            resolved_file_path = file_path
            resolved_sheet = sheet
            resolved_row = row
            resolved_columns = columns
            resolved_values = values
            resolved_updates = None
            resolved_header_row = header_row
            resolved_save_as = save_as

        if not isinstance(resolved_file_path, str) or not resolved_file_path.strip():
            raise ValueError("file_path must be a non-empty string")
        if resolved_updates is None:
            if resolved_sheet is None:
                raise ValueError("sheet is required")
            if resolved_row is None:
                raise ValueError("row is required")
            if resolved_columns is None:
                raise ValueError("columns is required")

        file_path = resolved_file_path
        sheet = resolved_sheet
        row = resolved_row
        columns = resolved_columns
        values = resolved_values
        updates_payload = resolved_updates
        header_row = resolved_header_row
        save_as = resolved_save_as

        excel_path = self._resolve_path(file_path)
        if not excel_path.exists() or not excel_path.is_file():
            raise ValueError("file_path does not exist")
        if excel_path.suffix.lower() not in {".xlsx", ".xlsm", ".xls"}:
            raise ValueError("file_path must be an Excel file (.xlsx/.xlsm/.xls)")

        if sheet is not None and not isinstance(sheet, (str, int)):
            raise ValueError("sheet must be a string or integer")
        if not isinstance(header_row, int) or header_row < 1:
            raise ValueError("header_row must be an integer >= 1")

        update_specs: list[tuple[str | int, int, list[str], list[Any]]] = []
        if updates_payload is None:
            if not isinstance(row, int) or row < 1:
                raise ValueError("row must be an integer >= 1")
            if not isinstance(columns, list) or not columns or any(not isinstance(col, str) or not col.strip() for col in columns):
                raise ValueError("columns must be a non-empty array of non-empty strings")

            if isinstance(values, list):
                if len(values) != len(columns):
                    raise ValueError("values array length must match columns length")
                mapped_values = values
            else:
                if len(columns) != 1:
                    raise ValueError("when values is a single value, columns must contain exactly one item")
                mapped_values = [values]
            if sheet is None:
                raise ValueError("sheet is required")
            update_specs.append((sheet, row, columns, mapped_values))
        else:
            if not isinstance(updates_payload, list) or not updates_payload:
                raise ValueError("updates must be a non-empty array when provided")
            for index, item in enumerate(updates_payload, start=1):
                if not isinstance(item, dict):
                    raise ValueError(f"updates[{index}] must be an object")

                item_sheet = item.get("sheet", sheet)
                item_row = item.get("row")
                item_columns = item.get("columns", columns)
                item_values = item.get("values")

                if not isinstance(item_sheet, (str, int)):
                    raise ValueError(f"updates[{index}].sheet must be a string or integer")
                if not isinstance(item_row, int) or item_row < 1:
                    raise ValueError(f"updates[{index}].row must be an integer >= 1")
                if not isinstance(item_columns, list) or not item_columns or any(not isinstance(col, str) or not col.strip() for col in item_columns):
                    raise ValueError(f"updates[{index}].columns must be a non-empty array of non-empty strings")

                if isinstance(item_values, list):
                    if len(item_values) != len(item_columns):
                        raise ValueError(f"updates[{index}].values length must match columns length")
                    normalized_values = item_values
                else:
                    if len(item_columns) != 1:
                        raise ValueError(f"updates[{index}] with scalar value must contain exactly one column")
                    normalized_values = [item_values]

                update_specs.append((item_sheet, item_row, item_columns, normalized_values))

        output_path = excel_path
        if isinstance(save_as, str) and save_as.strip():
            output_path = self._resolve_path(save_as)
            if output_path.suffix.lower() not in {".xlsx", ".xlsm", ".xls"}:
                raise ValueError("save_as must be an Excel file (.xlsx/.xlsm/.xls)")

        try:
            workbook = load_workbook(filename=str(excel_path))
        except Exception as exc:
            raise ValueError(f"Failed to open Excel file: {exc}") from exc

        worksheet_cache: dict[str, Any] = {}
        header_map_cache: dict[str, dict[str, int]] = {}

        def _resolve_worksheet(sheet_ref: str | int):
            if isinstance(sheet_ref, int):
                if sheet_ref < 0 or sheet_ref >= len(workbook.sheetnames):
                    raise ValueError(f"sheet index out of range: {sheet_ref}")
                sheet_name = workbook.sheetnames[sheet_ref]
            else:
                sheet_name = sheet_ref
                if sheet_name not in workbook.sheetnames:
                    raise ValueError(f"sheet not found: {sheet_name}")

            if sheet_name not in worksheet_cache:
                worksheet_cache[sheet_name] = workbook[sheet_name]
            return worksheet_cache[sheet_name]

        def _get_header_map(sheet_name: str, worksheet_obj: Any) -> dict[str, int]:
            if sheet_name not in header_map_cache:
                map_value: dict[str, int] = {}
                for col_idx in range(1, worksheet_obj.max_column + 1):
                    header_value = worksheet_obj.cell(row=header_row, column=col_idx).value
                    if isinstance(header_value, str) and header_value.strip():
                        map_value[header_value.strip()] = col_idx
                header_map_cache[sheet_name] = map_value
            return header_map_cache[sheet_name]

        updates: list[dict[str, Any]] = []
        rows_updated: list[int] = []
        sheets_updated: list[str] = []
        for target_sheet, target_row, target_columns, target_values in update_specs:
            worksheet = _resolve_worksheet(target_sheet)
            sheet_name = worksheet.title
            header_map = _get_header_map(sheet_name, worksheet)

            sheets_updated.append(sheet_name)
            rows_updated.append(target_row)
            for col_name, value in zip(target_columns, target_values):
                normalized_col = col_name.strip()
                if normalized_col not in header_map:
                    raise ValueError(f"column not found in header row for sheet '{sheet_name}': {normalized_col}")

                col_idx = header_map[normalized_col]
                worksheet.cell(row=target_row, column=col_idx).value = value
                updates.append(
                    {
                        "sheet_name": sheet_name,
                        "column": normalized_col,
                        "col_index": col_idx,
                        "row": target_row,
                        "value": self._to_json_safe(value),
                        "cell": f"{worksheet.cell(row=header_row, column=col_idx).column_letter}{target_row}",
                    }
                )

        output_path.parent.mkdir(parents=True, exist_ok=True)
        try:
            workbook.save(str(output_path))
        except Exception as exc:
            raise ValueError(f"Failed to save Excel file: {exc}") from exc

        return {
            "status": "success",
            "action": "update_sheet_row_values",
            "file_path": str(excel_path),
            "sheet_name": sheets_updated[0] if len(set(sheets_updated)) == 1 else None,
            "sheets_updated": sorted(set(sheets_updated)),
            "row": rows_updated[0] if len(rows_updated) == 1 else None,
            "rows_updated": sorted(set(rows_updated)),
            "header_row": header_row,
            "updates": updates,
            "update_count": len(updates),
            "save_as": str(output_path),
        }
